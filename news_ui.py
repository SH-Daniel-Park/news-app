# -*- coding: utf-8 -*-
import io
import json
import pandas as pd
import streamlit as st
import datetime as dt

from news_aggregator import (
    collect_articles,
    enrich_with_content,
    filter_by_publishers,
    extract_domain,
    DEFAULT_RSS_FEEDS,
)

st.set_page_config(page_title="📰 뉴스 키워드 수집/요약 대시보드", layout="wide")
st.title("📰 뉴스 키워드 수집/요약 대시보드")
st.caption("키워드로 여러 언론 기사를 모아보고, 본문/요약/키워드를 함께 확인하세요.")

# --------------------------- 사이드바 ---------------------------
with st.sidebar:
    st.header("검색 설정")
    query = st.text_input("검색 키워드", placeholder="예) 재정정책, 반도체, 환율 급등")
    max_results = st.slider("최대 기사 수", 10, 300, 60, step=10)
    newsapi_key = st.text_input("NewsAPI 키 (선택)", type="password")

    st.markdown("---")
    st.subheader("기간(선택)")
    use_date_range = st.toggle("기간 필터 사용", value=False)
    start_date = end_date = None
    if use_date_range:
        col1, col2 = st.columns(2)
        with col1:
            start_date = st.date_input("시작일", value=dt.date.today())
        with col2:
            end_date = st.date_input("종료일", value=dt.date.today())

    st.markdown("---")
    st.subheader("RSS 소스")
    use_default_rss = st.checkbox("샘플 기본 RSS 사용", True)
    uploaded_feeds = st.file_uploader("feeds.txt 업로드", type=["txt"])

    rss_feeds = []
    if use_default_rss:
        rss_feeds.extend(DEFAULT_RSS_FEEDS)
    if uploaded_feeds is not None:
        try:
            txt = uploaded_feeds.read().decode("utf-8", errors="ignore")
            for line in txt.splitlines():
                url = line.strip()
                if url and not url.startswith("#"):
                    rss_feeds.append(url)
        except Exception:
            st.warning("feeds.txt를 읽는 중 문제가 발생했습니다.")

    st.markdown("---")
    st.subheader("콘텐츠 처리")
    do_fetch_text = st.checkbox("기사 본문 수집 (크롤링)", True)
    do_summarize = st.checkbox("요약 생성", True)
    summary_len = st.slider("요약 문장 수", 2, 6, 3)
    do_keywords = st.checkbox("키워드(형태소) 추출", True)

    st.markdown("---")
    st.subheader("PDF 한글 폰트(선택) 업로드")
    pdf_font_file = st.file_uploader("NotoSansKR 등 .ttf/.otf 업로드하면 PDF 한글 표시가 좋습니다.", type=["ttf", "otf"])

    st.markdown("---")
    run = st.button("🔎 수집 시작", use_container_width=True)

# --------------------------- 유틸 ---------------------------
def _to_yyyymmdd(s: str) -> str:
    if not s or pd.isna(s):
        return ""
    try:
        d = pd.to_datetime(s, errors="coerce", utc=True)
        if pd.isna(d):
            d = pd.to_datetime(s, errors="coerce")
        if pd.isna(d):
            return ""
        return d.strftime("%Y-%m-%d")
    except Exception:
        return ""

def _truncate_kor(s: str, max_chars: int = 8) -> str:
    if s is None:
        return ""
    s = str(s)
    if len(s) <= max_chars:
        return s
    return s[:max_chars] + "…"

def _make_title_text(query, start_date, end_date, use_date_range=False):
    parts = [f"키워드: {query.strip()}"]
    if use_date_range and (start_date or end_date):
        s = start_date.strftime("%Y-%m-%d") if start_date else ""
        e = end_date.strftime("%Y-%m-%d") if end_date else ""
        parts.append(f"기간: {s} ~ {e}")
    return " / ".join([p for p in parts if p])

# --------------------------- 실행 ---------------------------
if run:
    if not query.strip():
        st.warning("키워드를 입력하세요.")
        st.stop()

    with st.spinner("기사를 수집하고 있습니다..."):
        raw = collect_articles(
            query=query,
            max_results=max_results,
            newsapi_key=newsapi_key.strip() or None,
            rss_feeds=rss_feeds if rss_feeds else None,
        )

    if not raw:
        st.info("관련 기사를 찾지 못했습니다.")
        st.stop()

    publishers = sorted(list({(it.get("publisher") or extract_domain(it.get("link","")) or "").strip()
                              for it in raw if it.get("link")}))
    with st.expander("언론사/도메인 필터"):
        allow = st.multiselect("포함할 언론사/도메인", options=publishers, default=[])

    filtered = filter_by_publishers(raw, allow_publishers=allow)

    # 기간 필터
    def parse_date_iso(s):
        try:
            return pd.to_datetime(s, utc=True).date()
        except Exception:
            return None

    if use_date_range and (start_date or end_date):
        if start_date and end_date and end_date < start_date:
            start_date, end_date = end_date, start_date
        tmp = []
        for it in filtered:
            d = parse_date_iso(it.get("published_at"))
            ok = True
            if start_date and d and d < start_date:
                ok = False
            if end_date and d and d > end_date:
                ok = False
            if ok:
                tmp.append(it)
        filtered = tmp

    if not filtered:
        st.info("필터 조건에 맞는 기사가 없습니다.")
        st.stop()

    # 본문/요약/키워드
    if do_fetch_text or do_summarize or do_keywords:
        with st.spinner("본문/요약/키워드를 생성 중입니다..."):
            enriched = enrich_with_content(
                filtered,
                do_fetch_text=do_fetch_text,
                do_summarize=do_summarize,
                do_keywords=do_keywords,
                summary_sentences=summary_len,
            )
    else:
        enriched = filtered

    df = pd.DataFrame(enriched)

    # 화면 표시는 URL 제외 + 날짜 YYYY-MM-DD 표준화
    df_display = df.copy()
    if "published_at" in df_display.columns:
        df_display["published_at"] = df_display["published_at"].map(_to_yyyymmdd)

    display_cols = ["title", "publisher", "published_at"]
    if do_summarize:
        display_cols.append("summary")
    if do_keywords:
        display_cols.append("keywords")

    st.success(f"총 {len(df_display)}건의 기사를 확보했습니다.")
    st.dataframe(df_display[display_cols], use_container_width=True, height=520)

    # 내부 링크 시리즈(엑셀/PDF에서 제목 하이퍼링크용)
    link_series = df["link"] if "link" in df.columns else pd.Series([""] * len(df))

    # 공통 타이틀 텍스트
    title_text = _make_title_text(query, start_date, end_date, use_date_range)

    # ---------------- Excel 다운로드 (첫 행 제목 + URL 제외 + 제목만 클릭 + 날짜 YYYY-MM-DD) ----------------
    from io import BytesIO

    df_excel = df_display[display_cols].copy()  # URL 제외, 날짜 이미 YYYY-MM-DD로 정규화

    engine = None
    try:
        import xlsxwriter
        engine = "xlsxwriter"
    except Exception:
        try:
            import openpyxl
            engine = "openpyxl"
        except Exception:
            engine = None

    if engine is None:
        st.error("xlsxwriter 또는 openpyxl이 필요합니다. requirements.txt에 추가 후 배포하세요.")
    else:
        output = BytesIO()
        if engine == "xlsxwriter":
            with pd.ExcelWriter(output, engine=engine) as writer:
                # 데이터는 2행부터 쓰기 (0-based로 startrow=1) → 1행(A1)은 타이틀
                df_excel.to_excel(writer, index=False, sheet_name="results", startrow=1)
                ws = writer.sheets["results"]
                cols = list(df_excel.columns)
                title_idx = cols.index("title") if "title" in cols else None

                # A1 ~ 마지막 컬럼 헤더까지 병합하여 타이틀 배치
                last_col_letter = chr(ord('A') + len(cols) - 1)
                title_format = writer.book.add_format({
                    "bold": True, "font_size": 12, "align": "left", "valign": "vcenter"
                })
                ws.merge_range(f"A1:{last_col_letter}1", title_text, title_format)

                # 제목만 클릭 가능 하이퍼링크 (데이터는 2행부터 시작이므로 r=2부터 보이게 → 0-based로 r=2-1=1 헤더, 데이터는 r>=2)
                # 실제 write_url은 0-based 좌표: 헤더는 row=1, 데이터 첫 행은 row=2
                # enumerate start=2 → Data starts at Excel row index 2 (0-based=1) but we wrote data startrow=1,
                # so hyperlinks must start from row=2 (0-based index=2) to align correctly?
                # Safer: iterate over df_excel with 0-based i and add +2 to row (title row 0, header row 1, data starts at 2).
                for i, (title, url) in enumerate(zip(df_excel["title"], link_series)):
                    if pd.notna(url) and str(url).strip().startswith("http"):
                        ws.write_url(2 + i, title_idx, str(url), string=str(title))
        else:
            # openpyxl path
            with pd.ExcelWriter(output, engine=engine) as writer:
                df_excel.to_excel(writer, index=False, sheet_name="results", startrow=1)
                ws = writer.sheets["results"]

                from openpyxl.styles import Font, Alignment
                cols = list(df_excel.columns)
                title_idx = cols.index("title") if "title" in cols else None

                # 병합 및 타이틀
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
                cell = ws.cell(row=1, column=1)
                cell.value = title_text
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal="left", vertical="center")

                # 제목 하이퍼링크 (데이터는 2행부터 → 헤더가 2행, 데이터는 3행부터)
                for i, (title, url) in enumerate(zip(df_excel["title"], link_series), start=3):
                    if pd.notna(url) and str(url).strip().startswith("http"):
                        c = ws.cell(row=i, column=title_idx + 1)
                        c.value = str(title)
                        c.hyperlink = str(url)
                        c.font = Font(color="0000EE", underline="single")

        st.download_button(
            "엑셀(.xlsx)로 다운로드",
            data=output.getvalue(),
            file_name=f"{query}_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    # -------------------------------------------------------------------

    # ---------------- PDF 다운로드 (첫 행 제목 + URL 제외 + 제목만 링크 + 날짜 YYYY-MM-DD) ---------
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except Exception:
        st.error("PDF 생성을 위해 reportlab 패키지가 필요합니다. requirements.txt에 reportlab을 추가하세요.")
    else:
        base_font_name = "Helvetica"
        if pdf_font_file is not None:
            try:
                font_bytes = pdf_font_file.read()
                font_name = "UserKoreanFont"
                import tempfile
                with tempfile.NamedTemporaryFile(delete=False, suffix="."+pdf_font_file.name.split(".")[-1]) as tf:
                    tf.write(font_bytes)
                    tmp_font_path = tf.name
                pdfmetrics.registerFont(TTFont(font_name, tmp_font_path))
                base_font_name = font_name
            except Exception:
                st.warning("업로드한 폰트를 등록하지 못했습니다. 기본 폰트로 진행합니다. (한글 표시가 깨질 수 있음)")

        styles = getSampleStyleSheet()
        title_style_link = ParagraphStyle(
            name="TitleLink",
            parent=styles["Normal"],
            fontName=base_font_name,
            fontSize=10,
            textColor=colors.HexColor("#1155cc"),
            leading=14,
        )
        cell_style = ParagraphStyle(
            name="Cell",
            parent=styles["Normal"],
            fontName=base_font_name,
            fontSize=9,
            leading=12,
        )
        header_style = ParagraphStyle(
            name="Header",
            parent=styles["Normal"],
            fontName=base_font_name,
            fontSize=12,
            leading=16,
            spaceAfter=6,
            textColor=colors.HexColor("#202124"),
        )

        # 상단 타이틀
        story = [Paragraph(title_text, header_style), Spacer(1, 6)]

        # 표 데이터 준비
        df_display = df.copy()
        if "published_at" in df_display.columns:
            df_display["published_at"] = df_display["published_at"].map(_to_yyyymmdd)

        rows = []
        header = ["제목", "언론사", "발행일"]
        rows.append(header)

        for idx, row in df_display.iterrows():
            title = str(row.get("title", ""))
            url = str(df.loc[idx].get("link", "")) if "link" in df.columns else ""
            pub_full = str(row.get("publisher", ""))
            pub_trunc = _truncate_kor(pub_full, 8)
            when = str(row.get("published_at", ""))  # YYYY-MM-DD

            if url.startswith("http"):
                title_para = Paragraph(f'<link href="{url}">{title}</link>', title_style_link)
            else:
                title_para = Paragraph(title, cell_style)

            pub_para = Paragraph(pub_trunc, cell_style)
            when_para = Paragraph(when, cell_style)

            rows.append([title_para, pub_para, when_para])

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=25, rightMargin=25, topMargin=20, bottomMargin=20)
        tbl = Table(rows, repeatRows=1, colWidths=[340, 90, 150])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#f1f3f4")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.HexColor("#202124")),
            ("FONTNAME", (0,0), (-1,-1), base_font_name),
            ("FONTSIZE", (0,0), (-1,0), 10),
            ("FONTSIZE", (0,1), (-1,-1), 9),
            ("ALIGN", (2,1), (2,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#fafafa")]),
            ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#dfe1e5")),
        ]))

        story.append(tbl)
        doc.build(story)
        pdf_bytes = buffer.getvalue()
        buffer.close()

        st.download_button(
            "PDF로 다운로드",
            data=pdf_bytes,
            file_name=f"{query}_results.pdf",
            mime="application/pdf",
            use_container_width=True,
        )
    # -------------------------------------------------------------------

else:
    st.info("좌측에서 키워드를 입력 후 수집 시작을 눌러주세요.")
